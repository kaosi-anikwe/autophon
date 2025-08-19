# ==============================================================================
# IMPORTS
# ==============================================================================

# Standard library imports
import os
import codecs
import shutil
import string
import random
import subprocess
import unicodedata
from datetime import datetime, timedelta, timezone
from app.utils.datetime_helpers import utc_now

# Third-party imports
import charset_normalizer
from praatio import textgrid
from dotenv import load_dotenv
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from praatio.utilities.constants import Interval
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Local imports
from app.extensions import db
from app.utils.logger import get_logger

# ==============================================================================
# CONFIGURATION
# ==============================================================================

load_dotenv()

UPLOADS = os.getenv("UPLOADS")
ADMIN = os.getenv("ADMIN")

logger = get_logger(__name__)

# ==============================================================================
# USER PROFILE & AUTH & ICON FUNCTIONS
# ==============================================================================


def generate_user_icon(name, user_id, force=False):
    """Generate user profile icon with initials"""
    image_path = os.path.join(UPLOADS, user_id, "profile.png")
    os.makedirs(os.path.dirname(image_path), exist_ok=True)

    if not os.path.exists(image_path) or force:
        # Load the template image
        template_path = os.path.join(ADMIN, "profile_template.png")
        image = Image.open(template_path).convert("RGBA")

        # Create a drawing context
        draw = ImageDraw.Draw(image)

        # Get the user's initials
        name_parts = name.split()
        if len(name_parts) >= 2:
            initials = f"{name_parts[0][:2]}{name_parts[1][:2]}"
        else:
            initials = name[:4] if len(name) >= 4 else name

        # Define the font size and type
        font_size = 100
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
        font = ImageFont.truetype(font_path, font_size)

        # Get the size of the image and the text
        image_width, image_height = image.size

        # Use textbbox instead of deprecated textsize
        bbox = draw.textbbox((0, 0), initials, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        # Calculate the position to center the text on the image
        x = (image_width - text_width) / 2 + 5
        y = (image_height - text_height) / 2 - 10

        # Draw the initials on the image
        draw.text((x, y), initials, font=font, fill="black")

        # Save to user folder
        image.save(image_path)

    return os.path.relpath(image_path)


# get captcha for register page
def generate_captcha() -> tuple:
    """
    Generates and returns a CAPTCHA image as well as the text for validation.
    """
    from captcha.image import ImageCaptcha

    image = ImageCaptcha()
    captcha_text = "".join(random.choices(string.digits, k=4))
    data = image.generate(captcha_text)

    return data, captcha_text


# ==============================================================================
# TEXT PROCESSING & ENCODING FUNCTIONS
# ==============================================================================


def missing_word_html(word_list, seperator=True) -> str:
    """Generate HTML for missing word display"""
    logger.debug("Generating missing word html")
    html_code = ""
    for i, word in enumerate(word_list, start=1):
        word_parts = word.split("\t")
        if len(word_parts) == 2:
            word_text = word_parts[0]
            phonemes = word_parts[1].split()
            colored_phonemes = []
            for phoneme in phonemes:
                if len(phoneme) > 1 and phoneme[-1].isdigit():
                    colored_phonemes.append(
                        f"<span class='text-secondary'>{phoneme}</span>"
                    )
                else:
                    colored_phonemes.append(
                        f"<span class='text-accent'>{phoneme}</span>"
                    )
            colored_phoneme_str = " ".join(colored_phonemes)
            n_digits = len(str(len(word_list))) - len(str(i))
            spaces = "".join(["&nbsp;"] * n_digits)
            if seperator:
                word_text = f"""<span><span class="after:content-['\\2502'] after:text-xl after:leading-3">{spaces}{i} </span> {word_text}</span>"""
            else:
                word_text = f"""<span>{spaces}{i} {word_text}</span>"""
            # number | word -> pro nun cia tion
            html_code += f"{word_text}\t{colored_phoneme_str}<br>\n"
        else:
            logger.info(f"Invalid entry: {word}")
    return html_code


def replace_decomposed(text):
    """Replace decomposed Unicode characters with precomposed equivalents"""
    # Normalize to NFD (decomposed form)
    nfd = unicodedata.normalize("NFD", text)

    result = []
    i = 0
    while i < len(nfd):
        if i + 1 < len(nfd) and unicodedata.combining(nfd[i + 1]):
            # Found a potentially decomposed character
            j = i + 1
            while j < len(nfd) and unicodedata.combining(nfd[j]):
                j += 1
            decomposed = nfd[i:j]
            precomposed = unicodedata.normalize("NFC", decomposed)

            if len(precomposed) == 1:
                # A precomposed version exists
                result.append(precomposed)
            else:
                # No precomposed version, keep decomposed form
                result.append(decomposed)
            i = j
        else:
            # Regular character
            result.append(nfd[i])
            i += 1

    return "".join(result)


def detect_encoding(file_path):
    """Detect file encoding using charset_normalizer"""
    with open(file_path, "rb") as file:
        raw_data = file.read()
        result = charset_normalizer.detect(raw_data)
        encoding = result["encoding"]
        logger.info(f"{file_path} encoding is: {encoding}")
        return encoding


def change_codec(file, codec="utf-8"):
    """Convert file encoding to specified codec"""
    # convert to input codec
    with codecs.open(file, "r", encoding=detect_encoding(file)) as file_in:
        text = file_in.read()
        decoded_text = text.encode(codec, errors="ignore")
    # write the utf-8 encoded text to the output file
    with open(file, "wb") as output_file:
        output_file.write(decoded_text)


def character_resolve(tg_path: str):
    """Remove composing decomposed characters from TextGrid file"""
    logger.info(f"Removing composing decomposed characters in file.")
    tg = textgrid.openTextgrid(tg_path, includeEmptyIntervals=True)
    for tier in tg.tiers:
        for start, stop, text in tier.entries:
            text = replace_decomposed(text)
            entry = Interval(start, stop, text)
            tier.insertEntry(
                entry, collisionMode="replace", collisionReportingMode="silence"
            )
    logger.info(f"Done composing decomposed characters in file.")
    tg.save(tg_path, format="long_textgrid", includeBlankSpaces=True)


# ==============================================================================
# EXCEL & PDF FUNCTIONS
# ==============================================================================


def get_font(font_size: int = 8, bold=None, gray=None) -> Font:
    """Get standardized font for Excel files"""
    if gray:
        font = Font(name="Josefin Sans", size=font_size, bold=bold, color="C0C0C0")
        return font
    font = Font(name="Josefin Sans", size=font_size, bold=bold)
    return font


def excel_to_pdf(input_file, output_file) -> bool:
    """Convert Excel file to PDF using LibreOffice"""
    # run these 2 commands first
    # make sure libreoffice and unconv are installed

    # ln -s /usr/lib/python3/dist-packages/uno.py /path/to/your/virtualenv/lib/python3.8/site-packages/uno.py
    # ln -s /usr/lib/python3/dist-packages/unohelper.py /path/to/your/virtualenv/lib/python3.8/site-packages/unohelper.py

    command = f"PYTHON=$(eval 'which python') UNOPATH=$(eval 'which libreoffice') /usr/bin/unoconv -f pdf -o {output_file} {input_file}"
    process = subprocess.run(command, shell=True, executable="/bin/bash")

    if process.returncode == 0:
        return True
    else:
        return False


def get_monthly_download(user_id, date, task_list: list, totals: dict):
    """Generate monthly download report as PDF"""
    # Load the Excel file
    workbook = load_workbook(f"{ADMIN}/monthly_download.xlsx")

    sheet = workbook.active

    # color fill
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    alignment = Alignment(horizontal="center", vertical="center")
    vertical_alignment = Alignment(vertical="center")

    # write date
    sheet["G5"].font = get_font()
    sheet["G5"] = date

    # write user_id
    sheet["A7"].font = get_font(9)
    sheet["A7"] = user_id

    # insert new rows for tasks
    num_of_rows = len(task_list)
    sheet.insert_rows(10, amount=num_of_rows)

    # write rows/tasks
    for i in range(num_of_rows):
        row = 10 + i
        if row % 2 == 1:  # change color of odd rows
            sheet[f"B{row}"].fill = gray_fill
            sheet[f"C{row}"].fill = gray_fill
            sheet[f"D{row}"].fill = gray_fill
            sheet[f"E{row}"].fill = gray_fill
            sheet[f"F{row}"].fill = gray_fill
            sheet[f"G{row}"].fill = gray_fill

        sheet[f"B{row}"].font = get_font()
        sheet[f"B{row}"] = task_list[i]["download_date"]
        sheet[f"B{row}"].alignment = vertical_alignment

        sheet[f"C{row}"].font = get_font()
        sheet[f"C{row}"] = task_list[i]["no_of_files"] or "N/A"
        sheet[f"C{row}"].alignment = alignment

        sheet[f"D{row}"].font = get_font()
        sheet[f"D{row}"] = task_list[i]["size"] or "N/A"
        sheet[f"D{row}"].alignment = alignment

        sheet[f"E{row}"].font = get_font()
        sheet[f"E{row}"] = (
            task_list[i]["lang"].replace("(suggested)", "").strip() or "N/A"
        )
        sheet[f"E{row}"].alignment = alignment

        sheet[f"F{row}"].font = get_font()
        sheet[f"F{row}"] = task_list[i]["words"] or "N/A"
        sheet[f"F{row}"].alignment = alignment

        sheet[f"G{row}"].font = get_font()
        sheet[f"G{row}"] = task_list[i]["task_status"]
        sheet[f"G{row}"].alignment = alignment

    # write total row
    total_row = 10 + num_of_rows + 1
    sheet[f"B{total_row}"].font = get_font(bold=True)
    sheet[f"B{total_row}"] = "Total usage"
    sheet[f"B{total_row}"].alignment = vertical_alignment

    sheet[f"C{total_row}"].font = get_font(bold=True)
    sheet[f"C{total_row}"] = totals["file_count"]
    sheet[f"C{total_row}"].alignment = alignment

    sheet[f"D{total_row}"].font = get_font(bold=True)
    sheet[f"D{total_row}"] = totals["total_size"]
    sheet[f"D{total_row}"].alignment = alignment

    sheet[f"E{total_row}"] = ""

    sheet[f"F{total_row}"].font = get_font(bold=True)
    sheet[f"F{total_row}"].alignment = alignment

    # write in totals box
    totals_box = total_row + 3
    # box dimensions
    min_col = sheet[f"C{totals_box}"].column
    min_row = sheet[f"C{totals_box}"].row
    max_col = sheet[f"F{totals_box + len(totals['lang_count'].keys()) -1}"].column
    max_row = sheet[f"F{totals_box + len(totals['lang_count'].keys()) -1}"].row

    for i, item in enumerate(sorted(totals["lang_count"].keys())):
        cur_index = totals_box + i
        # language
        sheet[f"E{cur_index}"].value = item
        sheet[f"E{cur_index}"].alignment = alignment
        # count
        sheet[f"F{cur_index}"].value = totals["lang_count"][item]
        sheet[f"F{cur_index}"].alignment = alignment
        if i != len(totals["lang_count"].keys()) + 1:
            sheet.insert_rows(cur_index + 1, amount=1)

    # draw border around totals box
    thin_border = Side(style="thin")
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = Border()
            # Apply top border to cells in the first row
            if cell.row == min_row:
                cell.border += Border(top=thin_border)
            # Apply bottom border to cells in the last row
            if cell.row == max_row:
                cell.border += Border(bottom=thin_border)
            # Apply left border to cells in the first column
            if cell.column == min_col:
                cell.border += Border(left=thin_border)
            # Apply right border to cells in the last column
            if cell.column == max_col:
                cell.border += Border(right=thin_border)

    # write footer
    footer_row = total_row + 2 + len(totals["lang_count"].keys()) + 2

    sheet[
        f"A{footer_row}"
    ] = "If you have any questions about this statement, please contact"
    sheet[f"A{footer_row}"].alignment = alignment
    sheet[f"A{footer_row}"].font = get_font()

    sheet[f"A{footer_row + 1}"] = "support@autophon.se"
    sheet[f"A{footer_row + 1}"].alignment = alignment
    sheet[f"A{footer_row + 1}"].font = get_font()

    # change font
    # get last row
    last_row = None
    for row in reversed(list(sheet.iter_rows())):
        for cell in row:
            if cell.value:
                last_row = cell.row
                break
        if last_row:
            break

    for row in sheet.iter_rows(
        min_row=10, max_row=last_row, min_col=3, max_col=sheet.max_column
    ):
        for cell in row:
            cell.font = get_font()

    # auto align
    for column in sheet.iter_cols(
        min_row=10, max_row=last_row, min_col=3, max_col=sheet.max_column
    ):
        sheet.column_dimensions[
            f"{get_column_letter(column[0].column)}"
        ].auto_size = True

    sheet.merge_cells(f"A{footer_row}:G{footer_row}")
    sheet.merge_cells(f"A{footer_row + 1}:G{footer_row + 1}")

    # set new logger.info area
    sheet.print_area = f"'{sheet.title}'!${get_column_letter(sheet.min_column)}${sheet.min_row}:${get_column_letter(sheet.max_column)}${sheet.max_row}"

    # Save the modified workbook
    output_path = f"{UPLOADS}/{user_id}/history"
    os.makedirs(output_path, exist_ok=True)
    file_name = (
        f"{str(datetime.strptime(f'{date}', '%b %Y').strftime('%y%m'))}_{user_id}"
    )
    output_file = f"{output_path}/{file_name}"
    workbook.save(f"{output_file}.xlsx")

    if not excel_to_pdf(f"{output_file}.xlsx", f"{output_file}.pdf"):
        raise Exception("There was a problem converting xlsx to pdf.")

    return output_file


# ==============================================================================
# SQLALCHEMY DATABASE FUNCTIONS
# ==============================================================================


def purge_unverified_accounts(timeframe: int = 48):
    """Delete unverified accounts from SQLAlchemy database after `timeframe` in hours\n
    Should be called withing an app context.
    """
    from app.models.user import User

    cutoff_time = datetime.now(timezone.utc) - timedelta(hours=timeframe)

    # Find unverified users older than timeframe
    unverified_users = User.query.filter(
        User.verified == False, User.created_at < cutoff_time
    ).all()

    deleted_count = 0
    for user in unverified_users:
        try:
            logger.info(f"Deleting unverified account: {user.email} (ID: {user.id})")
            delete_user_account(user.id)
            deleted_count += 1
        except Exception as e:
            logger.error(f"Failed to delete user {user.id}: {e}")

    logger.info(f"Purged {deleted_count} unverified accounts")
    return deleted_count


def delete_user_account(user_id: int):
    """Delete user account and associated data using SQLAlchemy"""
    from app.models.user import User
    from app.models.task import Task

    try:
        user = User.query.get(user_id)
        if not user:
            logger.warning(f"User {user_id} not found")
            return False

        # Mark user as deleted instead of hard delete for data integrity
        user.title = "deleted"
        user.first_name = "deleted"
        user.last_name = "deleted"
        user.email = f"deleted_{user_id}@deleted.com"
        user.deleted = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        user.update()

        # Delete associated tasks (cascade should handle this, but explicit is better)
        Task.query.filter_by(user_id=user_id).delete()

        # Delete user folders
        delete_folders(UPLOADS, str(user.uuid))

        db.session.commit()
        logger.info(f"Successfully deleted user account {user_id}")
        return True

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting user account {user_id}: {e}")
        return False


def populate_users(
    limit: datetime, include_deleted: bool = False, filename: str = None
):
    """Export users to Excel file using SQLAlchemy - matches original populate_users function signature"""
    from app.models.user import User

    new_filename = (
        os.path.join(ADMIN, "users", f"new_users_{limit.strftime('%y%m%d')}.xlsx")
        if not filename
        else filename
    )

    # Create Excel file
    new_workbook_path = create_users_excel_template(new_filename)
    new_workbook = load_workbook(new_workbook_path)
    new_sheet = new_workbook.active

    # Query users with date limit (registered/created_at <= limit)
    query = User.query.filter(User.created_at <= limit)

    # Apply deleted filter based on include_deleted parameter
    if not include_deleted:
        query = query.filter((User.deleted == None) | (User.deleted == ""))

    # Sort by registration date (created_at) ascending to match original
    users = query.order_by(User.created_at).all()

    # Add user data to Excel - maintaining original data structure
    for user in users:
        deleted_date = None
        if user.deleted and user.deleted != "":
            try:
                # Try to parse as datetime string and format for Excel
                deleted_date = datetime.strptime(
                    user.deleted, "%Y-%m-%d %H:%M:%S"
                ).strftime("%m/%d/%Y %H:%M:%S")
            except:
                # Fallback to original string if parsing fails
                deleted_date = user.deleted

        new_sheet.append(
            (
                user.uuid,  # Use UUID instead of old MongoDB id
                user.title or "",
                user.first_name or "",
                user.last_name or "",
                user.email or "",
                user.created_at.strftime(
                    "%m/%d/%Y %H:%M:%S"
                ),  # registered -> created_at
                user.verified,
                deleted_date,
                user.org or user.industry or "Non-Affiliated",
            )
        )

    new_workbook.save(new_workbook_path)
    logger.info(f"Exported {len(users)} users to {new_workbook_path}")
    return new_workbook_path


def populate_history(filename: str = None):
    """Export task history to Excel file using SQLAlchemy"""
    from app.models.task import Task
    from app.models.user import User

    new_filename = (
        os.path.join(
            ADMIN, "history", f"history_{datetime.now().strftime('%y%m%d')}.xlsx"
        )
        if not filename
        else filename
    )

    # Create Excel file
    new_workbook_path = create_history_excel_template(new_filename)
    new_workbook = load_workbook(new_workbook_path)
    new_sheet = new_workbook.active

    # Query all tasks with user information
    tasks = Task.query.join(User).order_by(Task.task_id).all()

    for task in tasks:
        new_sheet.append(
            (
                task.user_id,
                f"{task.owner.uuid}_{task.task_id}",
                task.task_id,
                task.download_date or task.created_at.strftime("%Y/%m/%d"),
                int((task.no_of_files or 0) / 2) or "",
                task.lang or (task.language.display_name if task.language else ""),
                task.size or 0,
                task.words or 0,
                task.task_status.value if task.task_status else "",
                task.deleted or "",
            )
        )

    new_workbook.save(new_workbook_path)
    logger.info(f"Exported {len(tasks)} tasks to {new_workbook_path}")
    return new_workbook_path


# ==============================================================================
# EXCEL TEMPLATE CREATION FUNCTIONS
# ==============================================================================


def create_users_excel_template(filename: str = None):
    """Create Excel template for user data export"""
    file_path = os.path.join(ADMIN, "users.xlsx") if not filename else filename
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    column_names = [
        "User ID",
        "Title",
        "First Name",
        "Last Name",
        "Email",
        "Date Created",
        "Verified",
        "Date Deleted",
        "Affiliation",
    ]

    # Add column headers
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num, value=column_name)

    # Ensure directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook.save(file_path)
    return str(file_path)


def create_history_excel_template(filename: str = None):
    """Create Excel template for task history export"""
    file_path = os.path.join(ADMIN, "history.xlsx") if not filename else filename
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Task History"

    column_names = [
        "user_id",
        "upload_id",
        "task_id",
        "date",
        "file_pairs",
        "language",
        "size",
        "words",
        "status",
        "deleted",
    ]

    # Add column headers
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num, value=column_name)

    # Ensure directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook.save(file_path)
    return str(file_path)


# ==============================================================================
# FILE SYSTEM UTILITIES
# ==============================================================================


def copy_file(src: str, dst: str):
    """Copy file from source to destination, creating directories as needed"""
    if os.path.exists(src):
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        shutil.copy(src, dst)


def delete_folders(folder_path: str, search_str: str) -> None:
    """Recursively delete all folders containing search string"""
    if os.path.exists(folder_path):
        for entry in os.scandir(folder_path):
            if entry.is_dir():
                if search_str in entry.path:
                    if os.path.exists(entry.path):
                        shutil.rmtree(entry.path)
                delete_folders(entry.path, search_str)
